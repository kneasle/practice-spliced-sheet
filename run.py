#!/usr/bin/env python3

from itertools import groupby
import sys
import re
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.fonts import Font
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Iterable, List, Tuple


FOOTER_TEXT = (
    "All compositions use near calls and are true."
    + "  C/Y/S, E/L, B/D and M/N can always be interchanged if falseness/less music is ok."
)

BELL_NAMES = "1234567890ETABCD"

STAGE = 8
TENOR = BELL_NAMES[STAGE - 1]
ROUNDS = BELL_NAMES[:STAGE]


def main():
    method_set = load_methods()
    touches = read_touches(sys.argv[1], method_set)
    touches.sort(key=lambda touch: (touch.length, -touch.runs))
    write_spreadsheet(method_set, touches, "Lincoln.xlsx")
    print(f"Written {len(touches)} touches")


###########
# METHODS #
###########


def load_methods() -> "MethodSet":
    return MethodSet(
        {
            # Core 7
            "C": Method("Cambridge", "-38-14-1258-36-14-58-16-78,12"),
            "Y": Method("Yorkshire", "-38-14-58-16-12-38-14-78,12"),
            "S": Method("Superlative", "-36-14-58-36-14-58-36-78,12"),
            "B": Method("Bristol", "-58-14.58-58.36.14-14.58-14-18,18"),
            "E": Method("Lessness", "-38-14-56-16-12-58-14-58,12"),
            "W": Method("Cornwall", "-56-14-56-38-14-58-14-58,18"),
            "L": Method("London", "38-38.14-12-38.14-14.58.16-16.58,12"),
            # Lincoln 11
            "N": Method("Double Norwich", "-14-36-58-18,18"),
            "V": Method("Deva", "-58-14.58-58.36-14-58-36-18,18"),
            "A": Method("Lancashire", "58-58.14-58-36-14-58.14-14.78,12"),
            "T": Method("Ytterbium", "-38-14-1256-16-12-58.16-12.78,12"),
            # Lincoln 15
            "D": Method("Double Coslany", "-14.58.36.14.58-18,18"),
            "M": Method("Mareham", "-58-14.58-12.38-12-18.36.12-18,18"),
            "G": Method("Glasgow", "36-56.14.58-58.36-14-38.16-16.38,18"),
            # "R": Method("Carolina Reaper", "38-38.18-56-18-34-18.16-16.78,12"),
        },
        groups=[("Core 7", 7), ("Friends", 7)],
        lines=[7, 11],
    )


class MethodSet:
    """A set of methods, including information about grouping and dividing lines"""

    def __init__(
        self,
        methods: Dict[str, "Method"],
        groups: List[Tuple[str, int]],
        lines: List[int],
    ) -> None:
        self.methods = methods
        self.groups = groups
        self.lines = lines


class Method:
    def __init__(self, name: str, pn_string: str) -> None:
        self.name: str = name

        BOB_PLACES = [1, 4]
        SINGLE_PLACES = [1, 2, 3, 4]

        # Parse the place notation
        pns = parse_pn(pn_string)

        # Use the place notation to generate the first lead of the method
        current_row = ROUNDS
        lead_rows = []
        for places in pns:
            lead_rows.append(current_row)
            current_row = transpose_row_by_pn(current_row, places)

        self.lead_rows: List[str] = lead_rows
        self.lead_head_plain: str = current_row
        self.lead_head_bob: str = transpose_row_by_pn(lead_rows[-1], BOB_PLACES)
        self.lead_head_single: str = transpose_row_by_pn(lead_rows[-1], SINGLE_PLACES)


###################
# LOADING TOUCHES #
###################


def read_touches(path: str, method_set: MethodSet) -> List["Touch"]:
    # I could do many things to make this regex more readable: like, for example, not using a regex.
    # Instead I will leave understanding this regex as a challenge for the reader.
    line_split_regex = re.compile(
        r"^\s*(?P<length>\d+)\s+(?P<calling>\S+)(\s+(?P<notes>\S.+?))?\s*$",
    )

    touches = []
    for line in open(path).read().splitlines(False):
        if line.lstrip().startswith("#"):
            continue

        re_match = line_split_regex.match(line)
        if re_match is None:
            print(f"Can't parse line {line.__repr__()}")
            exit(1)

        length = int(re_match.group("length"))
        call_string = re_match.group("calling")
        notes = re_match.group("notes")

        touches.append(Touch(length, call_string, notes, method_set.methods))

    return touches


class Touch:
    def __init__(
        self, length: int, call_string: str, notes: str, methods: Dict[str, Method]
    ) -> None:
        self.length = length
        self.call_string = call_string
        self.notes = notes

        # Parse the call string into a sequence of leads
        lead_regex = re.compile(r"(?P<method>[a-zA-Z])(?P<call>[*.])?")
        leads = [
            (match.group("method"), match.group("call"))
            for match in lead_regex.finditer(call_string)
        ]

        # Convert the lead sequence into a sequence of rows and a calling string
        rows, calls, last_lead_length = gen_rows_and_calls(call_string, leads, methods)

        # Determine which methods are rung
        self.method_counts = {}
        for shorthand, _ in leads:
            if shorthand in self.method_counts:
                self.method_counts[shorthand] += 1
            else:
                self.method_counts[shorthand] = 1
        # The last lead must be handled differently.  If we ring less than half a lead, then that
        # method should *only* be counted if it hasn't already been rung.  For example, both the
        # following should have one lead of Yorkshire:
        # - WS*LCE.BYY>     (Y is rung for 34 rows)
        # - CSCS.B*Y>       (Y is rung for 2 rows)
        # Basically, we want to not count a snap finish as a full lead but we *really* don't want to
        # have a method that's actually included (even for two rows) but doesn't show up in the
        # list.
        last_shorthand, _ = leads[-1]
        len_of_last_methods_lead = len(methods[last_shorthand].lead_rows)
        if last_lead_length < len_of_last_methods_lead / 2:
            # Finished in first half of the last lead, so reduce that method's lead count if needed
            self.method_counts[last_shorthand] = max(1, self.method_counts[last_shorthand] - 1)

        # Check that the given length was correct
        if self.length != len(rows):
            raise ValueError(
                f"{self.call_string} is given len {self.length} but has {len(rows)} rows"
            )
        # Count runs
        run_regex_front = re.compile("^(1234|2345|3456|4567|5678|4321|5432|6543|7654|8765).*$")
        run_regex_back = re.compile("^.*(1234|2345|3456|4567|5678|4321|5432|6543|7654|8765)$")
        self.runs = 0
        for row in rows:
            if run_regex_front.match(row):
                self.runs += 1
            if run_regex_back.match(row):
                self.runs += 1

        # Generate calling position string
        self.calling_position_string = ""
        for position, calls in groupby(calls, lambda call_pos: call_pos[1]):
            calls = [call for call, _ in calls]
            # Add the calls as efficiently as possible
            if all((call == "-" for call in calls)):
                # If all bobs, add nothing for one bob and a number for more than one
                if len(calls) > 1:
                    self.calling_position_string += str(len(calls))
            else:
                # If not all bobs, then just squash all the calls together
                for call in calls:
                    self.calling_position_string += call
            # Always add the position
            self.calling_position_string += position


def gen_rows_and_calls(
    call_string: str, leads, methods: Dict[str, Method]
) -> Tuple[List[str], List[Tuple[str, str]], int]:
    calls = []
    rows = []
    lead_head = ROUNDS
    last_lead_length = 0
    for method_shorthand, call_shorthand in leads:
        method = methods[method_shorthand]
        # Add the rows for this lead
        rows += [transpose_row_by_row(lead_head, row) for row in method.lead_rows]
        # Decide which lead head to go to
        if call_shorthand is None:
            lead_head = transpose_row_by_row(lead_head, method.lead_head_plain)
        elif call_shorthand == ".":
            lead_head = transpose_row_by_row(lead_head, method.lead_head_bob)
            calls.append(("-", calling_pos_at(lead_head, is_single=False)))
        elif call_shorthand == "*":
            lead_head = transpose_row_by_row(lead_head, method.lead_head_single)
            calls.append(("s", calling_pos_at(lead_head, is_single=True)))
        else:
            raise ValueError(f"Invalid call {call_shorthand}")
        last_lead_length = len(method.lead_rows)
    # Check for early rounds
    try:
        # Rounds always appears at the start, but snap finishes will make rounds appear again
        rounds_index = rows.index(ROUNDS, 1)
        # Remove the rows after rounds from the last lead
        last_lead_length -= len(rows) - rounds_index
        # Trim anything from rounds onwards
        rows = rows[:rounds_index]
    except ValueError:
        # Rounds doesn't appear early, so check that the comp comes round
        if lead_head != ROUNDS:
            print(f"{call_string} doesn't come round")
            assert False

    return (rows, calls, last_lead_length)


def calling_pos_at(row: str, is_single: bool = False) -> str:
    calling_positions = "LBTFVMWH" if is_single else "LIBFVMWH"
    return calling_positions[row.index(TENOR)]


#######################
# SPREADSHEET WRITING #
#######################


def write_spreadsheet(method_set: MethodSet, touches: List[Touch], path: str):
    # Header looks like:
    #       A/1    B/2    C/3   D/4    E/5     F/6      G/7     H/8        >=I/9
    #     +----+--------+-----+-----+------+--------+--------+-------+---------------
    #   1 |    |        |     |     |      |        |        |       |
    #     +----+--------+-----+-----+------+--------+--------+-------+---------------
    #     |    |                                                     |
    #   2 |    |                    <title>                          |
    #     |    |                                                     | <methods> ....
    #     +----+--------+-----+-----+------+--------+--------+-------+
    #   3 |    |    <made by me>    |      |        | back-  |       |
    #     +----+--------+-----+-----+ runs + queens + rounds + notes +---------------
    #   4 |    | length | <calling> |      |        |        |       | <method groups> ...
    #     +----+--------+-----+-----+------+--------+--------+-------+---------------
    #     |    |        |           |      |        |        |       |
    # >=5 |    |        |           |      |        |        |       |
    #                                ...
    #                             <touches>
    #                                ...
    #     |    |        |           |      |        |        |       |
    #     +----+--------+-----+-----+------+--------+--------+-------+---------------
    #     |                                   <note on calls>
    #     +----+--------+-----+-----+------+--------+--------+-------+---------------

    workbook = Workbook()
    sheet = workbook.active

    vertical_text = Alignment(text_rotation=90, horizontal="right")
    left_text = Alignment(horizontal="left", vertical="top")
    centre_text = Alignment(horizontal="center", vertical="top")
    right_text = Alignment(horizontal="right", vertical="top")

    FONT_FAMILY = "EB Garamond"
    FONT_SIZE = 10
    MATRIX_CELL_FILL = "FFCCCCCC"
    THIN_BORDER_COLOUR = "FF999999"

    THICK = Side(style="thick")
    NORMAL = Side(style="medium")
    THIN = Side(style="thin", color=THIN_BORDER_COLOUR)
    THICK_BOX = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

    num_methods = len(method_set.methods)
    methods_col = 1
    info_col = 1 + num_methods
    info_width = 5
    top_row = 1
    first_touch_row = top_row + 3

    length_col = info_col + 0
    notes_col = info_col + 1
    calling_col = info_col + 2
    runs_col = info_col + 4
    all_info_cols = [length_col, notes_col, calling_col, calling_col + 1, runs_col]

    # Set default font for all cells
    def set_col_font(col):
        for row in range(3 + len(touches) + 1):
            sheet.cell(top_row + row, col).font = Font(name=FONT_FAMILY, size=FONT_SIZE)

    for m_idx in range(num_methods):
        set_col_font(methods_col + m_idx)
    for i_idx in range(info_width):
        set_col_font(info_col + i_idx)

    # === TOP-LEFT CORNER ===
    # title
    sheet.merge_cells(
        start_column=info_col,
        start_row=top_row + 0,
        end_column=info_col + info_width - 1,
        end_row=top_row + 0,
    )
    title_cell = sheet.cell(top_row, info_col)
    title_cell.value = "50 Lincoln Touches"
    title_cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE * 4, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    # made by me
    sheet.merge_cells(
        start_column=info_col,
        start_row=top_row + 1,
        end_column=info_col + info_width - 1,
        end_row=top_row + 1,
    )
    made_by_me_cell = sheet.cell(top_row + 1, info_col + 0)
    made_by_me_cell.value = "Compiled by Ben White-Horne"
    made_by_me_cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE * 1.6, bold=True)
    made_by_me_cell.alignment = Alignment(horizontal="center", vertical="center")
    # merge calling header cells
    sheet.merge_cells(
        start_column=calling_col,
        start_row=top_row + 2,
        end_column=calling_col + 1,
        end_row=top_row + 2,
    )
    headers = [
        (length_col, "Length"),
        (calling_col, "Calling (* for single, . for bob, all near calls)"),
        (notes_col, "Notes"),
        (runs_col, "Runs"),
    ]
    for col, label in headers:
        cell = sheet.cell(top_row + 2, col)
        # 'Runs' and 'Length' look better right-aligned.  It seems that Google Sheets' text width
        # calculation isn't quite accurate, and for short words it's obvious that the text ends up
        # not exactly in the middle of the cell.
        cell.alignment = right_text if col in [length_col, runs_col] else centre_text
        cell.value = label

    # === METHODS ===
    for idx, shorthand in enumerate(method_set.methods):
        column = methods_col + idx
        method = method_set.methods[shorthand]
        # Determine the name
        name = " " + method.name
        if not method.name.startswith(shorthand):
            name += f" ({shorthand})"
        # Set the cell
        sheet.merge_cells(
            start_column=column,
            start_row=top_row,
            end_column=column,
            end_row=top_row + 1,
        )
        cell = sheet.cell(column=column, row=top_row)
        cell.value = name
        cell.alignment = vertical_text
    # Groups
    start_col = methods_col
    groups_row = top_row + 2
    for name, width in method_set.groups:
        sheet.merge_cells(
            start_column=start_col,
            start_row=groups_row,
            end_column=start_col + width - 1,
            end_row=groups_row,
        )
        cell = sheet.cell(groups_row, start_col)
        cell.value = name
        cell.alignment = centre_text
        start_col += width

    # === TOUCHES ===
    for idx, touch in enumerate(touches):
        row = first_touch_row + idx
        # Alignment (only the `calling` column is left-aligned)
        for col in all_info_cols:
            sheet.cell(row, col).alignment = left_text if col == calling_col else centre_text
        # Touch info
        for col in [calling_col, calling_col + 1]:
            sheet.cell(row, col).font = Font(name="Fira Code", size=FONT_SIZE)
        sheet.cell(row, length_col).value = touch.length
        sheet.cell(row, calling_col).value = touch.call_string
        sheet.cell(row, calling_col + 1).value = touch.calling_position_string
        sheet.cell(row, runs_col).value = touch.runs
        sheet.cell(row, notes_col).value = touch.notes
        # Method Matrix
        for meth_idx, shorthand in enumerate(method_set.methods):
            if shorthand in touch.method_counts:
                cell = sheet.cell(row, methods_col + meth_idx)
                if shorthand in touch.method_counts:
                    cell.fill = PatternFill(patternType="solid", fgColor=MATRIX_CELL_FILL)
                    cell.font = Font(name="Fira Code", size=FONT_SIZE)
                    if touch.method_counts[shorthand] > 1:
                        cell.alignment = centre_text
                        cell.value = touch.method_counts[shorthand]

    # === FOOTER ===
    footer_row = first_touch_row + len(touches)
    start_col = min(info_col, methods_col)
    # -1 to convert to an exclusive range
    end_col = max(info_col + info_width, methods_col + num_methods) - 1
    sheet.merge_cells(
        start_row=footer_row,
        start_column=start_col,
        end_row=footer_row,
        end_column=end_col,
    )
    sheet.cell(footer_row, start_col).value = FOOTER_TEXT
    sheet.cell(footer_row, start_col).border = Border(left=THICK, top=THICK, bottom=THICK)
    sheet.cell(footer_row, end_col).border = Border(right=THICK)

    # === ROW/COLUMN SIZES ===
    def get_col(col_idx):
        return sheet.column_dimensions[get_column_letter(col_idx)]

    def get_row(idx):
        return sheet.row_dimensions[top_row + idx]

    # Rows
    get_row(0).height = FONT_SIZE * 6  # Title
    get_row(1).height = FONT_SIZE * 4.5  # 'made by me'
    get_row(2).height = FONT_SIZE * 1.7  # 'made by me'
    for row in range(len(touches) + 1):
        get_row(3 + row).height = FONT_SIZE * 1.45
    # Columns
    vertical_text_column_width = 2.7
    get_col(length_col).width = 6.5
    get_col(runs_col).width = 5
    get_col(notes_col).width = max((len(t.notes or "") for t in touches)) * 0.95
    get_col(calling_col).width = max_len((t.call_string for t in touches)) * 1.3
    get_col(calling_col + 1).width = max_len((t.calling_position_string for t in touches)) * 1.3
    for method_idx in range(num_methods):
        col_name = get_column_letter(methods_col + method_idx)
        sheet.column_dimensions[col_name].width = vertical_text_column_width

    # === BORDERS ===
    sheet.cell(top_row, info_col).border = Border(top=THICK)
    for i in range(2):
        sheet.cell(top_row + i, info_col + info_width - 1).border = Border(right=THICK)
    # Column Headers
    for col in range(info_width):
        sheet.cell(top_row + 2, info_col + col).border = THICK_BOX
    # Methods box
    for method_idx in range(num_methods):
        i = methods_col + method_idx
        if method_idx == 0:
            left = THICK
        elif method_idx in method_set.lines:
            left = NORMAL
        else:
            left = None
        right = THICK if method_idx == num_methods - 1 else None
        # Set the borders
        sheet.cell(top_row, i).border = Border(
            left=left,
            top=THICK,
            right=right,
        )
        sheet.cell(top_row + 2, i).border = Border(
            left=left,
            right=right,
            top=THICK,
            bottom=THICK,
        )
    # Touches
    for touch_idx in range(len(touches)):
        row = first_touch_row + touch_idx
        # Determine what line is required under this cell
        if touch_idx == len(touches) - 1:
            bottom = THICK  # Thick border at the bottom of the box
        elif touch_idx % 5 == 4:
            bottom = NORMAL  # Thin line every 5 rows
        else:
            bottom = THIN
        # Info box
        for i in range(0, info_width):
            col = info_col + i
            sheet.cell(row, col).border = Border(
                left=THICK if col != calling_col + 1 else None,
                right=THICK if i == info_width - 1 else None,
                bottom=bottom,
            )
        # Method Matrix
        final_col = num_methods - 1
        for i in range(0, num_methods):
            if i == 0:
                left = THICK
            elif i in method_set.lines:
                left = NORMAL
            else:
                left = THIN
            sheet.cell(row, methods_col + i).border = Border(
                left=left,
                right=THICK if i == final_col else None,
                bottom=bottom,
            )

    workbook.save(path)


def max_len(strings: Iterable[str]) -> int:
    """Returns the maximum length of an `Iterable` of strings"""
    return max((len(s or "") for s in strings))


#######################
# PLACE NOTATION CODE #
#######################


# Place notation code is casually stolen from Wheatley:
# https://github.com/kneasle/wheatley/blob/9141bf8511dce737208731e55bfe138d48845319/wheatley/row_generation/helpers.py#L57


def transpose_row_by_row(lhs: str, rhs: str) -> str:
    return "".join((lhs[int(ch) - 1] for ch in rhs))


def transpose_row_by_pn(row: str, places: List[int]) -> str:
    new_row = ""
    index = 0
    while index < len(row):
        place = index + 1
        if place in places:
            # Don't do a swap
            new_row += row[index]
            index += 1
        else:
            assert place + 1 not in places
            # Swap two bells round
            new_row += row[index + 1]
            new_row += row[index]
            index += 2
    return new_row


def parse_pn(pn_str: str, expect_symmetric: bool = False) -> List[List[int]]:
    """Convert a place notation string into a list of places."""
    if "," in pn_str:
        pns = []
        for part in pn_str.split(","):
            pns += parse_pn(part, True)
        return pns

    if expect_symmetric:
        symmetric = not pn_str.startswith("+")
    else:
        symmetric = pn_str.startswith("&")

    # Assumes a valid place notation string is delimited by `.`
    # These can optionally be omitted around an `-` or `x`
    # We substitute to ensure `-` is surrounded by `.` and replace any `..` caused by `--` => `.-..-.
    dot_delimited_string = re.sub("[.]*[x-][.]*", ".-.", pn_str).strip(".&+ ")
    deduplicated_string = dot_delimited_string.replace("..", ".").split(".")

    # We suppress the type error here, because mypy will assign the list comprehension type 'List[object]',
    # not 'List[Places]'.
    converted: List[List[int]] = [
        [convert_bell_string(y) for y in place] if place != "-" else []  # type: ignore
        for place in deduplicated_string
    ]

    if symmetric:
        return converted + list(reversed(converted[:-1]))
    return converted


def convert_bell_string(bell: str) -> int:
    """Convert a single-char string representing a bell into an integer."""
    try:
        return BELL_NAMES.index(bell) + 1
    except ValueError as e:
        raise ValueError(f"'{bell}' is not known bell symbol") from e


if __name__ == "__main__":
    main()
